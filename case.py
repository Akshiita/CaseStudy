import time
import smtplib
import os
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# Set up Chrome WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome("path/to/chromedriver", options=chrome_options)

# Excel file path
excel_file = "path/to/excel/file.xlsx"

# LinkedIn login credentials
linkedin_username = "your_linkedin_username"
linkedin_password = "your_linkedin_password"

# Email notification settings
sender_email = "sender@gmail.com"
sender_password = "sender_password"
recipient_email = "recipient@gmail.com"
email_subject = "LinkedIn Notification Update"

def login_to_linkedin(username, password):
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)

    # Enter username
    username_input = driver.find_element(By.ID, "username")
    username_input.send_keys(username)

    # Enter password
    password_input = driver.find_element(By.ID, "password")
    password_input.send_keys(password)

    # Submit the login form
    submit_button = driver.find_element(By.XPATH, "//button[@type='submit']")
    submit_button.click()

    time.sleep(5)

def get_unread_counts():
    # Navigate to the LinkedIn notifications page
    driver.get("https://www.linkedin.com/notifications/")
    time.sleep(2)

    # Get the number of unread messages
    unread_messages_element = driver.find_element(By.CLASS_NAME, "msg-conversations-tab__badge-count")
    unread_messages = int(unread_messages_element.text.strip())

    # Get the number of unread notifications
    unread_notifications_element = driver.find_element(By.CLASS_NAME, "notifications-tab__badge-count")
    unread_notifications = int(unread_notifications_element.text.strip())

    return unread_messages, unread_notifications

def compare_with_previous_data(current_data, previous_data):
    comparison = {}
    for key, value in current_data.items():
        if key in previous_data:
            previous_value = previous_data[key]
            difference = value - previous_value
            comparison[key] = (value, difference)
        else:
            comparison[key] = (value, None)

    return comparison

def send_email_notification(sender, password, recipient, subject, body):
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender, password)
    server.sendmail(sender, recipient, msg.as_string())
    server.quit()

def update_excel_file(data):
    if not os.path.isfile(excel_file):
        workbook = openpyxl.Workbook()
        workbook.save(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    row = sheet.max_row + 1

    for col, value in enumerate(data.values(), start=1):
        sheet.cell(row=row, column=col, value=value)

    workbook.save(excel_file)

def read_previous_data():
    if not os.path.isfile(excel_file):
        return {}

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    data = {}

    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            data[row[0]] = row[1:]

    return data

def main():
    login_to_linkedin(linkedin_username, linkedin_password)
    unread_messages, unread_notifications = get_unread_counts()
    current_data = {
        "Unread Messages": unread_messages,
        "Unread Notifications": unread_notifications
    }
    previous_data = read_previous_data()

    comparison = compare_with_previous_data(current_data, previous_data)

    update_excel_file(current_data)

    email_body = f"""
    <html>
        <body>
            <h2>LinkedIn Notification Update</h2>
            <p>Number of unread messages: {current_data["Unread Messages"]}</p>
            <p>Number of unread notifications: {current_data["Unread Notifications"]}</p>
            <h3>Comparison with the previous occurrence:</h3>
            <ul>
                {"".join(f"<li>{key}: {value[0]} ({'+' if value[1] else ''}{value[1] or 0})</li>" for key, value in comparison.items())}
            </ul>
        </body>
    </html>
    """

    send_email_notification(sender_email, sender_password, recipient_email, email_subject, email_body)

    driver.quit()

if __name__ == "__main__":
    main()
